
"""
Utility functions for processing Excel files and extracting POR data.
"""

import logging

# Setup logging
logger = logging.getLogger(__name__)

import re
import os
from datetime import datetime, date, timezone
from typing import List, Dict, Any, Tuple, Optional
from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet
import json

def stringify(value: Any) -> str:
    """
    Convert value to string, handling dates specially.
    
    Args:
        value: Value to convert
        
    Returns:
        String representation of the value
    """
    if isinstance(value, (datetime, date)):
        return value.strftime("%d/%m/%Y")
    return str(value) if value is not None else ""


def to_float(value: Any) -> float:
    """
    Convert value to float, handling currency strings.
    
    Args:
        value: Value to convert
        
    Returns:
        Float value, 0.0 if conversion fails
    """
    if isinstance(value, (int, float)):
        return float(value)
    
    if isinstance(value, str):
        # Remove currency symbols and commas
        cleaned = re.sub(r'[£$,]', '', value.strip())
        try:
            return float(cleaned) if cleaned else 0.0
        except ValueError:
            return 0.0
    
    return 0.0

import xlrd # Added import for xlrd


def read_ws(stream) -> Tuple[List[List[Any]], Worksheet]:
    """
    Read Excel worksheet from stream, handling both .xls and .xlsx formats.
    """
    try:
        stream.seek(0)
        # Read the first few bytes to determine file type
        header = stream.read(8)
        stream.seek(0) # Reset stream position

        if header.startswith(b'PK'): # PK is the magic number for zip files (xlsx)
            return read_xlsx_file(stream)
        elif header.startswith(b'\xd0\xcf\x11\xe0\xa1\xb1\x1a\xe1'): # Magic number for OLE Compound Document (xls)
            return read_xls_file(stream)
        else:
            raise ValueError("Unsupported Excel file format. Please upload a .xls or .xlsx file.")
    except Exception as e:
        raise ValueError(f"Error reading Excel file: {str(e)}")

def read_xls_file(stream) -> Tuple[List[List[Any]], Worksheet]:
    """
    Read old Excel file (.xls) using xlrd.
    """
    try:
        # xlrd.open_workbook can take a file_contents argument
        # For BytesIO, we can pass the entire content
        workbook = xlrd.open_workbook(file_contents=stream.read())
        sheet = workbook.sheet_by_index(0) # Get the first sheet

        rows = []
        for row_idx in range(sheet.nrows):
            row_values = []
            for col_idx in range(sheet.ncols):
                cell_value = sheet.cell_value(row_idx, col_idx)
                row_values.append(cell_value)
            rows.append(row_values)

        # xlrd does not provide a 'worksheet' object directly like openpyxl
        # For compatibility, we might need to wrap it or adjust downstream code.
        # For now, we'll return the sheet object itself, assuming it has enough attributes.
        return rows, sheet # sheet object from xlrd

    except Exception as e:
        raise ValueError(f"Error reading .xls file: {str(e)}")

# Keep read_xlsx_file as is
def read_xlsx_file(stream) -> Tuple[List[List[Any]], Worksheet]:
    """
    Read modern Excel file (.xlsx) using openpyxl.
    """
    try:
        stream.seek(0)
        wb = load_workbook(stream, data_only=True, read_only=True)
        ws = wb.active

        if not ws:
            raise ValueError("No active worksheet found")

        # Convert to list for easier processing
        rows = [list(row) for row in ws.iter_rows(values_only=True)]

        return rows, ws

    except Exception as e:
        raise ValueError(f"Error reading Excel file: {str(e)}")


def find_vertical(rows: List[List[Any]], keyword: str) -> str:
    """
    Find value below keyword in vertical column.
    
    Args:
        rows: List of rows from worksheet
        keyword: Keyword to search for
        
    Returns:
        Found value as string, empty string if not found
    """
    if not rows:
        return ""
    
    keyword_lower = keyword.lower()
    max_cols = max(len(row) for row in rows)
    
    for col in range(max_cols):
        for row_idx, row in enumerate(rows):
            cell = row[col] if col < len(row) else None
            
            if isinstance(cell, str) and keyword_lower in cell.lower():
                # Look for value in next few rows
                for offset in range(1, min(5, len(rows) - row_idx)):
                    val = rows[row_idx + offset][col] if col < len(rows[row_idx + offset]) else None
                    if val not in (None, ""):
                        return stringify(val)
    
    return ""


def find_cell_by_keyword(ws: Worksheet, keyword: str, search_range: str = None) -> Tuple[int, int]:
    """
    Find cell position by keyword in worksheet.
    
    Args:
        ws: Worksheet to search
        keyword: Keyword to search for
        search_range: Optional range to limit search (e.g., "A1:Z50")
        
    Returns:
        Tuple of (row, column) where keyword was found, (-1, -1) if not found
    """
    try:
        if search_range:
            # Parse range like "A1:Z50"
            from openpyxl.utils import range_boundaries
            min_col, min_row, max_col, max_row = range_boundaries(search_range)
        else:
            min_col, min_row, max_col, max_row = 1, 1, ws.max_column, ws.max_row
        
        keyword_lower = keyword.lower()
        
        for row in range(min_row, max_row + 1):
            for col in range(min_col, max_col + 1):
                cell_value = ws.cell(row=row, column=col).value
                if isinstance(cell_value, str) and keyword_lower in cell_value.lower():
                    return row, col
        
        return -1, -1
        
    except Exception:
        return -1, -1


def extract_por_data_by_map(ws: Worksheet) -> Dict[str, Any]:
    """
    Extract POR data according to the exact parsing map provided.
    Following the exact cell references from the POR Parsing Map.
    
    Args:
        ws: Worksheet to process
        
    Returns:
        Dictionary containing extracted POR data
    """
    data = {}
    
    try:
        # SHOW PRICE (Y/N) (F28) = F29
        show_price = ws.cell(row=29, column=6).value  # F29
        data['show_price'] = stringify(show_price) if show_price else ''
        
        # DATE ORDER RAISED (A1) = A2
        date_order_raised = ws.cell(row=2, column=1).value  # A2
        data['date_order_raised'] = stringify(date_order_raised) if date_order_raised else ''
        
        # DATE REQUIRED BY (H1) = H2
        date_required_by = ws.cell(row=2, column=8).value  # H2
        data['date_required_by'] = stringify(date_required_by) if date_required_by else ''
        
        # SHIP / PROJECT NAME (B1) = B2
        ship_project_name = ws.cell(row=2, column=2).value  # B2
        data['ship_project_name'] = capitalize_text(ship_project_name) if ship_project_name else ''
        
        # SUPPLIER (if known) (D1) = D2
        supplier = ws.cell(row=2, column=4).value  # D2
        data['supplier'] = capitalize_text(supplier) if supplier else ''
        
        # SPECIFICATION / STANDARDS / CERTIFICATION / INSPECTION REQUIREMENTS / INVOICE NOTE A28 = A29
        specification_standards = ws.cell(row=29, column=1).value  # A29
        data['specification_standards'] = capitalize_text(specification_standards) if specification_standards else ''
        
        # Supplier Contact Name (A33) = C33
        supplier_contact_name = ws.cell(row=33, column=3).value  # C33
        data['supplier_contact_name'] = capitalize_text(supplier_contact_name) if supplier_contact_name else ''
        
        # Specific Supplier Contact Email (A34) = C34
        supplier_contact_email = ws.cell(row=34, column=3).value  # C34
        data['supplier_contact_email'] = supplier_contact_email if supplier_contact_email else ''
        
        # Quote Ref: (If Known) (A35) = C35
        quote_ref = ws.cell(row=35, column=3).value  # C35
        data['quote_ref'] = capitalize_text(quote_ref) if quote_ref else ''
        
        # Quote date: (If Known) (A36) = C36
        quote_date = ws.cell(row=36, column=3).value  # C36
        data['quote_date'] = stringify(quote_date) if quote_date else ''
        
        # REQUESTOR NAME (F33) = F34
        requestor_name = ws.cell(row=34, column=6).value  # F34
        if not requestor_name:
            # Fallback: search for "REQUESTOR" keyword
            requestor_pos = find_cell_by_keyword(ws, "REQUESTOR")
            if requestor_pos[0] != -1:
                requestor_name = ws.cell(row=requestor_pos[0] + 1, column=requestor_pos[1]).value
        data['requestor_name'] = capitalize_text(requestor_name) if requestor_name else ''
        
        # Add debug logging
        print(f"DEBUG: Extracted POR data: {data}")
        
        return data
        
    except Exception as e:
        print(f"Error extracting POR data: {str(e)}")
        import traceback
        traceback.print_exc()
        return data


def extract_line_items_by_map(ws: Worksheet) -> List[Dict[str, Any]]:
    """
    Extract line items according to the exact parsing map provided.
    
    Args:
        ws: Worksheet to process
        
    Returns:
        List of line item dictionaries
    """
    items = []
    
    try:
        # Get supplier information for business rule application
        supplier = ws.cell(row=2, column=4).value  # D2 - SUPPLIER (if known)
        supplier_str = str(supplier).lower().strip() if supplier else ""
        
        # According to map: rows 6-26 for line items
        start_row = 6
        end_row = min(26, ws.max_row)
        
        for row in range(start_row, end_row + 1):
            try:
                # JOB / CONTRACT No. (A4) = A6 to A26
                job_contract = ws.cell(row=row, column=1).value  # Column A
                
                # OP No. (B4) = B6 to B26
                op_no = ws.cell(row=row, column=2).value  # Column B
                
                # MATERIAL / SERVICE DESCRIPTION (C4 & C5) = C6 to C26
                description = ws.cell(row=row, column=3).value  # Column C
                
                # QUANTITY (G4) = G6 to G26
                quantity = ws.cell(row=row, column=7).value  # Column G
                
                # PRICE EACH £ (H4) = H6 to H26
                price_each = ws.cell(row=row, column=8).value  # Column H
                
                # LINE TOTAL (I4) = I6 to I26
                line_total = ws.cell(row=row, column=9).value  # Column I
                
                # Check for end of line items (ORDER TOTAL)
                if isinstance(description, str) and "ORDER TOTAL" in description.upper():
                    break
                
                # Apply business rule: EAPL supplier with IWO or Supply and Fit
                if description and isinstance(description, str):
                    desc_lower = description.lower()
                    
                    # Check if supplier is EAPL and description contains relevant keywords
                    if (supplier_str == "eapl" and 
                        ("iwo" in desc_lower or "supply and fit" in desc_lower) and
                        "work i.w.o" in desc_lower):
                        
                        # Replace "Work I.W.O" with "Provide Labour I.W.O"
                        description = description.replace("Work I.W.O", "Provide Labour I.W.O")
                        description = description.replace("work i.w.o", "Provide Labour I.W.O")
                        description = description.replace("WORK I.W.O", "Provide Labour I.W.O")
                        
                        logger.info(f"Applied EAPL business rule: Modified description for row {row}")
                
                # Only add if we have at least a description or job/contract
                if description or job_contract:
                    items.append({
                        "job": job_contract,
                        "op": op_no,
                        "desc": description,
                        "qty": quantity,
                        "price": price_each,
                        "ltot": line_total
                    })
                
            except Exception as e:
                print(f"Error processing row {row}: {str(e)}")
                continue
        
        return items
        
    except Exception as e:
        print(f"Error extracting line items: {str(e)}")
        return []


def get_order_total_by_map(ws: Worksheet) -> float:
    """
    Find order total according to the parsing map.
    Reads directly from cell H29 (row 29, column 8).
    
    Args:
        ws: Worksheet to search
        
    Returns:
        Order total as float, 0.0 if not found
    """
    try:
        # Read directly from cell H29 (row 29, column 8)
        cell_value = ws.cell(row=29, column=8).value
        
        if cell_value is not None:
            # Convert to float, handling different formats
            if isinstance(cell_value, (int, float)):
                return float(cell_value)
            elif isinstance(cell_value, str):
                # Handle currency strings and commas
                cleaned = re.sub(r'[£$,]', '', str(cell_value).strip())
                cleaned = re.sub(r',', '', cleaned)
                try:
                    return float(cleaned)
                except ValueError:
                    pass
        
        # Fallback: Look for "ORDER TOTAL" in the description column (C) around row 26
        for row in range(20, 30):  # Search around expected location
            try:
                cell_value = ws.cell(row=row, column=3).value  # Column C
                if isinstance(cell_value, str) and "ORDER TOTAL" in cell_value.upper():
                    # Look for numeric value in the same row or next few rows
                    for col in range(1, ws.max_column + 1):
                        val = ws.cell(row=row, column=col).value
                        if isinstance(val, (int, float)) and val > 0:
                            return float(val)
                    
                    # Check next few rows if not found in same row
                    for offset in range(1, 5):
                        try:
                            for col in range(1, ws.max_column + 1):
                                val = ws.cell(row=row + offset, column=col).value
                                if isinstance(val, (int, float)) and val > 0:
                                    return float(val)
                        except IndexError:
                            break
            except IndexError:
                continue
        
        return 0.0
        
    except Exception:
        return 0.0


def validate_excel_structure(rows: List[List[Any]]) -> bool:
    """
    Validate that Excel file has expected structure.
    
    Args:
        rows: List of rows from worksheet
        
    Returns:
        True if structure is valid
    """
    if not rows or len(rows) < 5:
        return False
    
    # Check for common POR keywords
    keywords = ['requestor', 'material', 'quantity', 'price', 'total']
    found_keywords = 0
    
    for row in rows[:10]:  # Check first 10 rows
        for cell in row:
            if isinstance(cell, str):
                cell_lower = cell.lower()
                for keyword in keywords:
                    if keyword in cell_lower:
                        found_keywords += 1
                        break
    
    return found_keywords >= 3  # At least 3 keywords should be found

def allowed_file(filename: str) -> bool:
    """
    Check if file is allowed for attachments.
    ACCEPTS ALL FILE TYPES - no restrictions for maximum compatibility.
    This ensures drag-and-drop from Outlook and all file types work.
    """
    # Accept everything - no restrictions for file attachments
    return True


def allowed_excel_file(filename: str) -> bool:
    """
    Check if file extension is allowed for main POR uploader.
    Accepts ALL Excel formats including legacy and variations.
    """
    if not filename or not isinstance(filename, str):
        return False
    
    # If no extension, reject (POR files must have extensions)
    if '.' not in filename:
        return False
    
    # Get extension
    extension = filename.rsplit('.', 1)[1].lower()
    
    # Comprehensive Excel format support including legacy
    excel_extensions = {
        # Modern Excel formats
        'xlsx', 'xlsm', 'xlsb', 'xltx', 'xltm',
        # Legacy Excel formats
        'xls', 'xlt', 'xlm', 'xla', 'xlw',
        # OpenOffice/LibreOffice Calc formats (often compatible)
        'ods', 'ots',
        # CSV formats (often used with Excel)
        'csv', 'tsv',
        # Other Excel-compatible formats
        'xml', 'mhtml', 'mht'
    }
    
    return extension in excel_extensions


def get_file_type_icon(file_type: str) -> str:
    """Get the appropriate icon for a file type."""
    icon_map = {
        'quote': '💰',
        'original': '📄',
        'correspondence': '📝',
        'updates': '🔄',
        'por': '📋',
        'other': '📎'
    }
    return icon_map.get(file_type, '📎')


def capitalize_text(value):
    """Capitalize text values, handling None and non-string values."""
    if value is None:
        return None
    if isinstance(value, str):
        return value.upper()
    return value

def clean_query_string(query_string):
    """Custom filter to clean query string by removing from_search parameter."""
    if not query_string:
        return ''
    
    # Parse the query string
    from urllib.parse import parse_qs, urlencode
    params = parse_qs(query_string)
    
    # Remove the from_search parameter if it exists
    if 'from_search' in params:
        del params['from_search']
    
    # Rebuild the query string
    return urlencode(params, doseq=True)


def get_company_info(company: str) -> dict:
    """Get company configuration information."""
    from company_config import get_company_config
    return get_company_config(company)


def add_change_to_history(por, field, old_value, new_value):
    """Add a change to the POR's change history."""
    try:
        # Parse existing history
        history = json.loads(por.change_history) if por.change_history else []
        
        # Create change record
        change = {
            'field': field,
            'old_value': old_value,
            'new_value': new_value,
            'timestamp': datetime.now(timezone.utc).isoformat()
        }
        
        # Add to history
        history.append(change)
        
        # Limit to 100 changes
        if len(history) > 100:
            history = history[-100:]
        
        # Update POR
        por.change_history = json.dumps(history)
        por.current_change_index = len(history) - 1
        
        return True
    except Exception as e:
        # logger.error(f"Error adding change to history: {str(e)}")
        return False


def undo_last_change(por):
    """Undo the last change for a POR."""
    try:
        # Parse existing history
        history = json.loads(por.change_history) if por.change_history else []
        
        if not history or por.current_change_index < 0:
            return False, "No changes to undo"
        
        # Get the last change
        change = history[por.current_change_index]
        
        # Revert the field
        setattr(por, change['field'], change['old_value'])
        
        # Move index back
        por.current_change_index -= 1
        
        return True, f"Undid change to {change['field']}"
        
    except Exception as e:
        logger.error(f"Error undoing change: {str(e)}")
        return False, str(e)


def redo_last_change(por):
    """Redo the last undone change for a POR."""
    try:
        # Parse existing history
        history = json.loads(por.change_history) if por.change_history else []
        
        if not history or por.current_change_index >= len(history) - 1:
            return False, "No changes to redo"
        
        # Get the next change
        change = history[por.current_change_index + 1]
        
        # Apply the change
        setattr(por, change['field'], change['new_value'])
        
        # Move index forward
        por.current_change_index += 1
        
        return True, f"Redid change to {change['field']}"
        
    except Exception as e:
        # logger.error(f"Error redoing change: {str(e)}")
        return False, str(e)

def detect_content_type_from_line_items(line_items: list, por_description: str = "") -> str:
    """Detect content type based on line item descriptions using rule-based pattern matching."""
    
    # Combine all line item descriptions for comprehensive analysis
    all_descriptions = []
    for item in line_items:
        desc = item.get('desc', '') or ''
        if desc:
            all_descriptions.append(desc)
    
    # Also include POR description if available
    if por_description:
        all_descriptions.append(por_description)
    
    if all_descriptions:
        # Join all descriptions for comprehensive analysis
        combined_text = ' '.join(all_descriptions)
        
        # Use rule-based pattern matching for content type detection
        logger.info("📝 Using rule-based pattern matching for content type detection")
    
    # Work IWO patterns (highest priority)
    work_iwo_patterns = [
        r'work\s+iwo',
        r'provide\s+labour\s+iwo',
        r'labour\s+iwo',
        r'iwo\s+work',
        r'work\s+only',
        r'labour\s+only',
        r'provide\s+labour',
        r'labour\s+services',
        r'work\s+services',
        r'installation\s+work',
        r'fitting\s+work',
        r'repair\s+work',
        r'maintenance\s+work'
    ]
    
    # Supply and Fit patterns
    supply_fit_patterns = [
        r'supply\s+and\s+fit',
        r'supply\s*&\s*fit',
        r'fit\s+and\s+supply',
        r'install\s+and\s+supply',
        r'supply\s+install',
        r'install\s+supply',
        r'fitting\s+and\s+supply',
        r'supply\s+and\s+installation',
        r'installation\s+and\s+supply',
        r'fit\s+supply',
        r'supply\s+fit',
        r'install\s+supply',
        r'supply\s+install'
    ]
    
    def fuzzy_match(text, patterns):
        """Fuzzy matching for content type detection."""
        if not text:
            return False
        
        text_lower = text.lower()
        
        for pattern in patterns:
            if re.search(pattern, text_lower, re.IGNORECASE):
                return True
        
        return False
    
    # Check all line items for patterns
    for item in line_items:
        description = item.get('desc', "") or ""
        
        # Check for Work IWO first (highest priority)
        if fuzzy_match(description, work_iwo_patterns):
            return 'work_iwo'
        
        # Check for Supply and Fit
        if fuzzy_match(description, supply_fit_patterns):
            return 'supply_and_fit'
    
    # Check POR description as fallback
    if por_description:
        if fuzzy_match(por_description, work_iwo_patterns):
            return 'work_iwo'
        if fuzzy_match(por_description, supply_fit_patterns):
            return 'supply_and_fit'
    
    # Default to supply if no patterns found
    return 'supply'
